import pandas as pd
import io
from openpyxl import load_workbook
import xlsxwriter

class ExcelHandler:
    @staticmethod
    def read_excel(file_buffer):
        """엑셀 파일을 읽어서 DataFrame으로 반환"""
        try:
            # 여러 시트가 있을 수 있으므로 모든 시트 읽기
            excel_data = pd.read_excel(file_buffer, sheet_name=None)
            return excel_data
        except Exception as e:
            raise Exception(f"엑셀 파일 읽기 오류: {str(e)}")
    
    @staticmethod
    def dataframe_to_excel(dataframes_dict):
        """DataFrame 딕셔너리를 엑셀 파일로 변환"""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            for sheet_name, df in dataframes_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def get_sheet_names(file_buffer):
        """엑셀 파일의 시트 이름들을 반환"""
        try:
            workbook = load_workbook(file_buffer)
            return workbook.sheetnames
        except Exception as e:
            return []
